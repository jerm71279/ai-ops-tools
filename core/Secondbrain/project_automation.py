#!/usr/bin/env python3
"""
Project Automation for Network Installs
Watches ECS for new install projects and automatically:
1. Assigns next available IP block from registry
2. Generates Network Checklist from template
3. Updates ECS project notes with IP scheme
4. Populates Engineering SOW tab from Pricing SOW equipment list

Run daily via cron or manually when needed.
"""
import os
import sys
import json
import requests
import re
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

# Try to import openpyxl for SOW updates
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed - SOW updates will be skipped")

load_dotenv()


class ProjectAutomation:
    def __init__(self):
        self.tenant_id = os.getenv('AZURE_TENANT_ID')
        self.client_id = os.getenv('AZURE_CLIENT_ID')
        self.client_secret = os.getenv('AZURE_CLIENT_SECRET')
        self.base_url = "https://graph.microsoft.com/v1.0"
        self.token = None
        self.site_id = None
        self.projects_list_id = 'a310c2d1-634f-44d0-862b-6750bf8788ce'
        self.tasks_list_id = '92c85920-31bd-49ae-a4ae-8cee71ee5f39'
        self.tickets_list_id = 'd135fb01-4f30-429d-94ad-b908a0c6a9f7'

        # Paths
        self.base_dir = Path(__file__).parent
        self.ip_registry_path = self.base_dir / "data" / "customer_ip_registry.json"
        self.checklist_template = self.base_dir / "output" / "Network_Installation_Checklist_Template.md"
        self.output_dir = self.base_dir / "output"
        self.processed_file = self.base_dir / "data" / "processed_projects.json"
        self.sharepoint_local = self.base_dir / "input_documents" / "sharepoint_all" / "oberaconnect_technical" / "Documents"

        # Ensure directories exist
        self.output_dir.mkdir(exist_ok=True)
        (self.base_dir / "data").mkdir(exist_ok=True)

        # Keywords that identify network install projects
        self.install_keywords = [
            'network install',
            'network installation',
            'new install',
            'site install',
            'network deployment',
            'network build',
            'network setup'
        ]

    def get_token(self):
        """Get OAuth token"""
        if self.token:
            return self.token

        url = f"https://login.microsoftonline.com/{self.tenant_id}/oauth2/v2.0/token"
        data = {
            'grant_type': 'client_credentials',
            'client_id': self.client_id,
            'client_secret': self.client_secret,
            'scope': 'https://graph.microsoft.com/.default'
        }
        response = requests.post(url, data=data)
        if response.status_code != 200:
            raise Exception(f"Failed to get token: {response.text}")

        self.token = response.json().get('access_token')
        return self.token

    def get_headers(self):
        """Get auth headers"""
        return {
            'Authorization': f'Bearer {self.get_token()}',
            'Content-Type': 'application/json'
        }

    def find_site(self):
        """Find EE TEAM / SOC TEAM site"""
        resp = requests.get(
            f"{self.base_url}/sites?search=SOCTEAM",
            headers=self.get_headers()
        )
        for site in resp.json().get('value', []):
            if 'SOCTEAM' in site.get('webUrl', ''):
                self.site_id = site['id']
                return True
        return False

    def load_ip_registry(self):
        """Load IP registry"""
        if self.ip_registry_path.exists():
            with open(self.ip_registry_path) as f:
                return json.load(f)
        return {
            "schema": {
                "base_network": "10.55.0.0/16",
                "block_size": 4,
                "gateway_offset": 1,
                "description": "Each customer gets 4 x /24 blocks - first is native VLAN, rest reserved"
            },
            "customers": [],
            "next_available_block": 0,
            "last_updated": None
        }

    def save_ip_registry(self, registry):
        """Save IP registry"""
        registry['last_updated'] = datetime.now().isoformat()
        with open(self.ip_registry_path, 'w') as f:
            json.dump(registry, f, indent=2)

    def load_processed_projects(self):
        """Load list of already processed project IDs"""
        if self.processed_file.exists():
            with open(self.processed_file) as f:
                return json.load(f)
        return {"processed": []}

    def save_processed_projects(self, processed):
        """Save processed projects list"""
        with open(self.processed_file, 'w') as f:
            json.dump(processed, f, indent=2)

    def assign_ip_block(self, customer_name, customer_id):
        """Assign next available IP block to customer"""
        registry = self.load_ip_registry()

        # Check if customer already has assignment
        for customer in registry['customers']:
            if customer['id'] == customer_id:
                return customer  # Already assigned

        # Get next block
        next_block = registry['next_available_block']
        base_third_octet = next_block * 4  # Each customer gets 4 /24s

        if base_third_octet > 252:
            raise Exception("IP registry full - no more blocks available")

        # Create assignment
        assignment = {
            "id": customer_id,
            "name": customer_name,
            "assigned_blocks": [
                f"10.55.{base_third_octet}.0/24",
                f"10.55.{base_third_octet + 1}.0/24",
                f"10.55.{base_third_octet + 2}.0/24",
                f"10.55.{base_third_octet + 3}.0/24"
            ],
            "active_block": f"10.55.{base_third_octet}.0/24",
            "default_gateway": f"10.55.{base_third_octet}.1",
            "reserved_for_expansion": [
                f"10.55.{base_third_octet + 1}.0/24",
                f"10.55.{base_third_octet + 2}.0/24",
                f"10.55.{base_third_octet + 3}.0/24"
            ],
            "dns": ["8.8.8.8", "8.8.4.4"],
            "status": "assigned",
            "assigned_date": datetime.now().strftime("%Y-%m-%d")
        }

        registry['customers'].append(assignment)
        registry['next_available_block'] = next_block + 1
        self.save_ip_registry(registry)

        return assignment

    def generate_checklist(self, project_name, customer_name, ip_assignment, project_info=None):
        """Generate Network Checklist from template"""
        if not self.checklist_template.exists():
            print(f"  Template not found: {self.checklist_template}")
            return None

        template = self.checklist_template.read_text()

        # Create header with project info
        date_str = datetime.now().strftime("%Y-%m-%d")
        safe_name = re.sub(r'[^\w\s-]', '', customer_name).strip().replace(' ', '_')

        header = f"""# Network Installation Checklist
## Customer: {customer_name}

**Date Created:** {date_str}
**Project:** {project_name}
**Site Address:** {project_info.get('address', 'TBD') if project_info else 'TBD'}
**Contact:** {project_info.get('contact', 'TBD') if project_info else 'TBD'}

---

## IP Addressing Scheme

| Block | Purpose | Gateway |
|-------|---------|---------|
| **{ip_assignment['active_block']}** | Native VLAN (Active) | {ip_assignment['default_gateway']} |
| {ip_assignment['reserved_for_expansion'][0]} | Reserved for expansion | - |
| {ip_assignment['reserved_for_expansion'][1]} | Reserved for expansion | - |
| {ip_assignment['reserved_for_expansion'][2]} | Reserved for expansion | - |

**DNS:** {', '.join(ip_assignment['dns'])}

---

"""
        # Remove the original header from template and add our custom one
        # Find where Phase 1 starts
        phase1_start = template.find("## Phase 1")
        if phase1_start > 0:
            checklist_content = header + template[phase1_start:]
        else:
            checklist_content = header + template

        # Save checklist locally
        output_file = self.output_dir / f"Network_Checklist_{safe_name}_{date_str}.md"
        output_file.write_text(checklist_content)

        return output_file, checklist_content

    def find_support_site(self):
        """Find the Support site where customer folders are located"""
        resp = requests.get(
            f"{self.base_url}/sites?search=support",
            headers=self.get_headers()
        )
        for site in resp.json().get('value', []):
            if 'support' in site.get('webUrl', '').lower():
                return site['id']
        return None

    def upload_checklist_to_sharepoint(self, customer_name, checklist_content, filename):
        """Upload checklist to customer folder in SharePoint Support site"""
        # Find the Support site (where customer folders are located)
        support_site_id = self.find_support_site()

        if not support_site_id:
            print("    WARNING: Could not find Support site")
            return None

        # Get the Support site's Documents drive
        resp = requests.get(
            f"{self.base_url}/sites/{support_site_id}/drives",
            headers=self.get_headers()
        )

        drives = resp.json().get('value', [])
        drive_id = None
        for drive in drives:
            if drive.get('name') == 'Documents':
                drive_id = drive['id']
                break

        if not drive_id:
            print("    WARNING: Could not find Documents drive in Support site")
            return None

        # Create customer folder path - upload to Technical Docs subfolder
        # Keep spaces in folder name to match SharePoint folder naming convention
        safe_customer = re.sub(r'[<>:"/\\|?*]', '', customer_name).strip()
        folder_path = f"{safe_customer}/Technical Docs"

        # Upload file to customer's Technical Docs folder
        headers = {
            'Authorization': f'Bearer {self.get_token()}',
            'Content-Type': 'text/plain'
        }

        upload_url = f"{self.base_url}/drives/{drive_id}/root:/{folder_path}/{filename}:/content"

        resp = requests.put(
            upload_url,
            headers=headers,
            data=checklist_content.encode('utf-8')
        )

        if resp.status_code in [200, 201]:
            web_url = resp.json().get('webUrl', '')
            return web_url
        else:
            print(f"    WARNING: Upload failed: {resp.status_code} - {resp.text[:100]}")
            return None

    def upload_checklist_to_soc_team(self, customer_name, checklist_content, filename):
        """Upload checklist to SOC TEAM site Projects folder (attached to ECS project)"""
        # Get the SOC TEAM site's Documents drive
        resp = requests.get(
            f"{self.base_url}/sites/{self.site_id}/drives",
            headers=self.get_headers()
        )

        drives = resp.json().get('value', [])
        drive_id = None
        for drive in drives:
            if drive.get('name') == 'Documents':
                drive_id = drive['id']
                break

        if not drive_id:
            print("    WARNING: Could not find Documents drive in SOC TEAM site")
            return None

        # Create project folder path
        safe_customer = re.sub(r'[<>:"/\\|?*]', '', customer_name).strip()
        folder_path = f"Projects/{safe_customer}"

        # Upload file
        headers = {
            'Authorization': f'Bearer {self.get_token()}',
            'Content-Type': 'text/plain'
        }

        upload_url = f"{self.base_url}/drives/{drive_id}/root:/{folder_path}/{filename}:/content"

        resp = requests.put(
            upload_url,
            headers=headers,
            data=checklist_content.encode('utf-8')
        )

        if resp.status_code in [200, 201]:
            web_url = resp.json().get('webUrl', '')
            return web_url
        else:
            print(f"    WARNING: SOC TEAM upload failed: {resp.status_code} - {resp.text[:100]}")
            return None

    def update_project_notes(self, project_id, ip_assignment, checklist_url=None, support_sharepoint_url=None):
        """Update ECS project with IP scheme in notes"""
        notes = f"""IP Scheme Assigned: {datetime.now().strftime("%Y-%m-%d")}
Native VLAN: {ip_assignment['active_block']}
Gateway: {ip_assignment['default_gateway']}
Reserved: {', '.join(ip_assignment['reserved_for_expansion'])}
DNS: {', '.join(ip_assignment['dns'])}"""

        if support_sharepoint_url:
            notes += f"\nSharePoint: {support_sharepoint_url}"

        if checklist_url:
            notes += f"\nChecklist: {checklist_url}"

        update_data = {
            "fields": {
                "Notes": notes
            }
        }

        resp = requests.patch(
            f"{self.base_url}/sites/{self.site_id}/lists/{self.projects_list_id}/items/{project_id}",
            headers=self.get_headers(),
            json=update_data
        )

        return resp.status_code == 200

    def create_project_task(self, project_id, customer_name, assignee=None, due_date=None):
        """Create a task in Engineering Tasks for the Kanban board"""
        # Generate short customer code
        words = customer_name.split()
        if len(words) >= 2:
            code = ''.join(w[0].upper() for w in words[:3])
        else:
            code = customer_name[:4].upper()

        task_data = {
            'fields': {
                'TaskTitle': f'{code} - Network Discovery',
                'TaskDescription': f'Perform network discovery for {customer_name}. Document current network topology, devices, and infrastructure.',
                'ProjectID': str(project_id),
                'Status': 'Not Started',  # Kanban: Not Started -> Scheduled -> In Progress -> Complete
                'Priority': 'Medium',
                'AssignedTo': assignee or 'Unassigned',
                'Phase': 'Discovery',
                'EstimatedHours': 4
            }
        }

        if due_date:
            task_data['fields']['DueDate'] = due_date

        resp = requests.post(
            f"{self.base_url}/sites/{self.site_id}/lists/{self.tasks_list_id}/items",
            headers=self.get_headers(),
            json=task_data
        )

        if resp.status_code in [200, 201]:
            return resp.json().get('id')
        else:
            print(f"    WARNING: Failed to create task: {resp.status_code}")
            return None

    def update_project_status(self, project_id, status='In Progress', phase='Implementation'):
        """Update project status in ECS"""
        update_data = {
            'fields': {
                'Status': status,
                'ProjectPhase': phase
            }
        }

        resp = requests.patch(
            f"{self.base_url}/sites/{self.site_id}/lists/{self.projects_list_id}/items/{project_id}",
            headers=self.get_headers(),
            json=update_data
        )

        return resp.status_code == 200

    def create_project_ticket(self, project_id, customer_name, project_name, assignee=None):
        """Create a ticket linked to project for tracking through lifecycle"""
        ticket_data = {
            'fields': {
                'TicketTitle': f'{customer_name} - Project Ticket',
                'Description': f'Master ticket for tracking {project_name} through project lifecycle. All tasks and work items are linked via ProjectID.',
                'ProjectID': str(project_id),
                'Customer': customer_name,
                'Status': 'Open',
                'Priority': 'Medium',
                'TicketType': 'Project',
                'Source': 'Automation',
                'AssignedTo': assignee or 'Unassigned'
            }
        }

        resp = requests.post(
            f"{self.base_url}/sites/{self.site_id}/lists/{self.tickets_list_id}/items",
            headers=self.get_headers(),
            json=ticket_data
        )

        if resp.status_code in [200, 201]:
            ticket_id = resp.json().get('id')
            # Update the task with the ticket ID for linking
            return ticket_id
        else:
            print(f"    WARNING: Failed to create ticket: {resp.status_code} - {resp.text[:100]}")
            return None

    def link_task_to_ticket(self, task_id, ticket_id):
        """Link a task to a ticket via TicketID field"""
        update_data = {
            'fields': {
                'TicketID': str(ticket_id)
            }
        }

        resp = requests.patch(
            f"{self.base_url}/sites/{self.site_id}/lists/{self.tasks_list_id}/items/{task_id}",
            headers=self.get_headers(),
            json=update_data
        )

        return resp.status_code == 200

    def update_project_ticket_id(self, project_id, ticket_id):
        """Update project with linked ticket ID"""
        update_data = {
            'fields': {
                'TicketID': str(ticket_id)
            }
        }

        resp = requests.patch(
            f"{self.base_url}/sites/{self.site_id}/lists/{self.projects_list_id}/items/{project_id}",
            headers=self.get_headers(),
            json=update_data
        )

        return resp.status_code == 200

    # ==================== SOW Functions ====================

    def find_customer_folder(self, customer_name):
        """Find customer folder in local SharePoint sync"""
        if not self.sharepoint_local.exists():
            return None

        # Try exact match first
        for folder in self.sharepoint_local.iterdir():
            if folder.is_dir() and folder.name.lower() == customer_name.lower():
                return folder

        # Try partial match
        for folder in self.sharepoint_local.iterdir():
            if folder.is_dir() and customer_name.lower() in folder.name.lower():
                return folder

        return None

    def find_sow_file(self, customer_folder):
        """Find SOW xlsx file in customer folder"""
        sow_folder = customer_folder / "Statements of Work (SOWs)"
        if not sow_folder.exists():
            return None

        # Look for main SOW file (starts with 'SOW-' not containing 'Camera')
        for f in sow_folder.glob("SOW-*.xlsx"):
            if 'camera' not in f.name.lower():
                return f

        # Fallback: any SOW file not containing 'Camera'
        for f in sow_folder.glob("*.xlsx"):
            if 'SOW' in f.name.upper() and 'camera' not in f.name.lower():
                return f

        return None

    def extract_equipment_from_pricing(self, wb):
        """Extract equipment list from Pricing SOW sheet"""
        equipment = {'locations': {}}

        if 'Pricing SOW' not in wb.sheetnames:
            return equipment

        ws = wb['Pricing SOW']
        current_location = 'Main'

        for row in ws.iter_rows(min_row=2, max_row=50, max_col=5):
            desc = row[0].value
            mfr = row[1].value
            model = row[2].value
            qty = row[4].value

            if desc:
                desc_str = str(desc).strip()

                # Skip formula results, totals, and empty/placeholder rows
                if desc_str.startswith('=') or 'total' in desc_str.lower() or 'labor' in desc_str.lower() or 'e/m' in desc_str.lower():
                    continue
                if desc_str == 'Description' or desc_str == 'Manufacturer':
                    continue

                # Check if this is a location header
                if any(kw in desc_str.lower() for kw in ['office', 'road', 'location', 'site', 'building']):
                    if not mfr and not model:
                        current_location = desc_str
                        equipment['locations'][current_location] = []
                        continue

                # Add equipment item (only if has manufacturer/model)
                mfr_str = str(mfr).strip() if mfr else ''
                model_str = str(model).strip() if model else ''
                if mfr_str.startswith('=') or model_str.startswith('='):
                    continue

                if mfr_str or model_str:
                    if current_location not in equipment['locations']:
                        equipment['locations'][current_location] = []

                    equipment['locations'][current_location].append({
                        'description': desc_str,
                        'manufacturer': mfr_str,
                        'model': model_str,
                        'quantity': int(qty) if qty and str(qty).isdigit() else 1
                    })

        return equipment

    def generate_engineering_sow_summary(self, equipment, customer_name):
        """Generate Engineering SOW summary text"""
        lines = []
        lines.append(f"Network Infrastructure Installation for {customer_name}")
        lines.append("")

        for location, items in equipment['locations'].items():
            if not items:
                continue

            lines.append(f"{location}:")

            # Group by type
            firewalls = [i for i in items if 'firewall' in i['description'].lower() or 'sonicwall' in i['manufacturer'].lower()]
            switches = [i for i in items if 'switch' in i['description'].lower()]
            aps = [i for i in items if 'ap' in i['description'].lower() or 'u7' in i['model'].lower() or 'access point' in i['description'].lower()]
            cabling = [i for i in items if 'cat' in i['description'].lower() or 'drop' in i['description'].lower()]
            other = [i for i in items if i not in firewalls + switches + aps + cabling]

            if firewalls:
                for item in firewalls:
                    lines.append(f"  - Install and configure {item['manufacturer']} {item['model']} firewall")
                    lines.append(f"    - Set up firewall rules and security policies")
                    lines.append(f"    - Configure VPN (if required)")

            if switches:
                for item in switches:
                    qty = item['quantity']
                    lines.append(f"  - Install {qty}x {item['manufacturer']} {item['model']} switch{'es' if qty > 1 else ''}")
                    lines.append(f"    - Configure ports and PoE")

            if aps:
                total_aps = sum(i['quantity'] for i in aps)
                lines.append(f"  - Install {total_aps}x UniFi access point{'s' if total_aps > 1 else ''}")
                lines.append(f"    - Configure SSIDs and wireless security")
                lines.append(f"    - Optimize placement for coverage")

            if cabling:
                total_drops = sum(i['quantity'] for i in cabling)
                lines.append(f"  - Install {total_drops}x CAT6 drop{'s' if total_drops > 1 else ''}")

            for item in other:
                if item['description'] and item['description'] not in ['', ' ']:
                    lines.append(f"  - {item['description']}")

            lines.append("")

        # Calculate engineering hours
        num_sites = len([loc for loc, items in equipment['locations'].items() if items])
        num_firewalls = sum(
            len([i for i in items if 'firewall' in i['description'].lower() or 'sonicwall' in i['manufacturer'].lower()])
            for items in equipment['locations'].values()
        )
        has_other_devices = any(
            any(i for i in items if 'firewall' not in i['description'].lower() and 'sonicwall' not in i['manufacturer'].lower())
            for items in equipment['locations'].values()
        )

        discovery_hours = num_sites  # 1 hour per site
        firewall_hours = num_firewalls  # 1 hour per firewall
        other_device_hours = 1 if has_other_devices else 0  # 1 hour for all other devices combined
        onsite_hours = num_sites  # 1 hour per site for verification

        total_hours = discovery_hours + firewall_hours + other_device_hours + onsite_hours

        lines.append("Engineering Hours Estimate:")
        lines.append(f"  - Network Discovery: {discovery_hours} hour{'s' if discovery_hours != 1 else ''} ({num_sites} site{'s' if num_sites != 1 else ''})")
        if num_firewalls > 0:
            lines.append(f"  - Firewall Configuration: {firewall_hours} hour{'s' if firewall_hours != 1 else ''} ({num_firewalls} firewall{'s' if num_firewalls != 1 else ''})")
        if has_other_devices:
            lines.append(f"  - Other Device Configuration: {other_device_hours} hour")
        lines.append(f"  - Onsite Install/Verification: {onsite_hours} hour{'s' if onsite_hours != 1 else ''}")
        lines.append(f"  - Total: {total_hours} hours")
        lines.append("")

        # Add standard deliverables
        lines.append("Deliverables:")
        lines.append("  - Network documentation and diagrams")
        lines.append("  - Device configuration backups")
        lines.append("  - Credentials documentation (secure)")

        return '\n'.join(lines)

    def update_engineering_sow_tab(self, sow_file, summary_text):
        """Update the Engineering SOW tab with the summary"""
        if not HAS_OPENPYXL:
            return False

        wb = openpyxl.load_workbook(sow_file)

        if 'Engineering SOW' not in wb.sheetnames:
            print("    Warning: Engineering SOW sheet not found")
            return False

        ws = wb['Engineering SOW']

        # B13 is the merged cell for Scope of Work content
        ws['B13'] = summary_text

        wb.save(sow_file)
        return True

    def update_customer_sow(self, customer_name):
        """Main function to update customer's Engineering SOW tab"""
        if not HAS_OPENPYXL:
            print("    Skipping SOW update (openpyxl not installed)")
            return False

        # Find customer folder
        customer_folder = self.find_customer_folder(customer_name)
        if not customer_folder:
            print(f"    Warning: Customer folder not found for '{customer_name}'")
            return False

        # Find SOW file
        sow_file = self.find_sow_file(customer_folder)
        if not sow_file:
            print(f"    Warning: SOW file not found in {customer_folder}")
            return False

        print(f"    Found SOW: {sow_file.name}")

        # Extract equipment from Pricing SOW
        wb = openpyxl.load_workbook(sow_file)
        equipment = self.extract_equipment_from_pricing(wb)

        if not equipment['locations']:
            print(f"    Warning: No equipment found in Pricing SOW")
            return False

        # Generate summary
        summary = self.generate_engineering_sow_summary(equipment, customer_name)

        # Update Engineering SOW tab
        if self.update_engineering_sow_tab(sow_file, summary):
            print(f"    Updated Engineering SOW tab")

            # Upload updated file to SharePoint
            upload_success = self.upload_sow_to_sharepoint(customer_name, sow_file)
            if upload_success:
                print(f"    Uploaded SOW to SharePoint")
            return True
        else:
            print(f"    Warning: Failed to update Engineering SOW tab")
            return False

    def upload_sow_to_sharepoint(self, customer_name, sow_file):
        """Upload updated SOW file to SharePoint"""
        # Find OberaConnect Technical site
        resp = requests.get(
            f"{self.base_url}/sites?search=OberaConnect Technical",
            headers=self.get_headers()
        )

        tech_site_id = None
        for site in resp.json().get('value', []):
            if 'technical' in site.get('displayName', '').lower():
                tech_site_id = site['id']
                break

        if not tech_site_id:
            return False

        # Get Documents drive
        resp = requests.get(
            f"{self.base_url}/sites/{tech_site_id}/drives",
            headers=self.get_headers()
        )

        drive_id = None
        for drive in resp.json().get('value', []):
            if drive.get('name') == 'Documents':
                drive_id = drive['id']
                break

        if not drive_id:
            return False

        # Upload file
        folder_path = f"{customer_name}/Statements of Work (SOWs)"
        headers = {
            'Authorization': f'Bearer {self.get_token()}',
            'Content-Type': 'application/octet-stream'
        }

        with open(sow_file, 'rb') as f:
            content = f.read()

        upload_url = f"{self.base_url}/sites/{tech_site_id}/drives/{drive_id}/root:/{folder_path}/{sow_file.name}:/content"
        resp = requests.put(upload_url, headers=headers, data=content)

        return resp.status_code in [200, 201]

    # ==================== End SOW Functions ====================

    def is_install_project(self, project_name, description=""):
        """Check if project is a network install based on keywords"""
        text = f"{project_name} {description}".lower()

        # Check for install keywords
        for keyword in self.install_keywords:
            if keyword in text:
                return True

        # Also check for customer names with common install patterns
        # e.g., "Customer Name - Network" or "Customer Name Network Install"
        if 'network' in text and any(word in text for word in ['new', 'install', 'deploy', 'build', 'setup']):
            return True

        return False

    def get_projects(self):
        """Get all projects from ECS"""
        resp = requests.get(
            f"{self.base_url}/sites/{self.site_id}/lists/{self.projects_list_id}/items?expand=fields",
            headers=self.get_headers()
        )
        return resp.json().get('value', [])

    def process_new_projects(self):
        """Main function - find and process new install projects"""
        print("=" * 60)
        print("Project Automation - Network Installs")
        print("=" * 60)
        print(f"Started: {datetime.now().isoformat()}")
        print()

        if not self.find_site():
            print("ERROR: Could not find SharePoint site")
            return

        print(f"Connected to SharePoint site")

        # Load processed projects
        processed = self.load_processed_projects()
        processed_ids = set(processed.get('processed', []))

        # Get all projects
        projects = self.get_projects()
        print(f"Found {len(projects)} total projects")

        new_installs = []

        for project in projects:
            project_id = project.get('id')
            fields = project.get('fields', {})
            project_name = fields.get('ProjectName', fields.get('Title', ''))
            description = fields.get('Description', '')
            customer = fields.get('Customer', project_name)
            notes = fields.get('Notes', '')
            status = fields.get('Status', '')

            # Skip already processed
            if str(project_id) in processed_ids:
                continue

            # Skip completed projects
            if status and status.lower() in ['completed', 'cancelled', 'on hold']:
                continue

            # Check if already has IP assignment
            if 'IP Scheme' in notes or '10.55.' in notes:
                processed_ids.add(str(project_id))
                continue

            # Check if this is an install project
            if self.is_install_project(project_name, description):
                new_installs.append({
                    'id': project_id,
                    'name': project_name,
                    'customer': customer,
                    'description': description,
                    'fields': fields
                })

        print(f"Found {len(new_installs)} new install projects to process")
        print()

        for project in new_installs:
            print(f"Processing: {project['name']}")

            # Create customer ID from name
            customer_id = re.sub(r'[^\w]', '-', project['customer'].lower()).strip('-')
            customer_id = re.sub(r'-+', '-', customer_id)

            try:
                # 1. Assign IP block
                ip_assignment = self.assign_ip_block(project['customer'], customer_id)
                print(f"  Assigned IP: {ip_assignment['active_block']}")

                # 2. Generate checklist
                checklist_result = self.generate_checklist(
                    project_name=project['name'],
                    customer_name=project['customer'],
                    ip_assignment=ip_assignment
                )
                checklist_path = None
                checklist_content = None
                if checklist_result:
                    checklist_path, checklist_content = checklist_result
                    print(f"  Generated checklist: {checklist_path.name}")

                # 3. Upload checklist to SharePoint Support site (customer Technical Docs)
                support_sharepoint_url = None
                if checklist_content:
                    support_sharepoint_url = self.upload_checklist_to_sharepoint(
                        customer_name=project['customer'],
                        checklist_content=checklist_content,
                        filename=checklist_path.name
                    )
                    if support_sharepoint_url:
                        print(f"  Uploaded to Support site: {support_sharepoint_url[:60]}...")
                    else:
                        print(f"  WARNING: Support site upload failed")

                # 4. Upload checklist to SOC TEAM site (attached to project)
                checklist_url = None
                if checklist_content:
                    checklist_url = self.upload_checklist_to_soc_team(
                        customer_name=project['customer'],
                        checklist_content=checklist_content,
                        filename=checklist_path.name
                    )
                    if checklist_url:
                        print(f"  Attached to project: {checklist_url[:60]}...")
                    else:
                        print(f"  WARNING: SOC TEAM upload failed")

                # 5. Update ECS project notes with links
                if self.update_project_notes(project['id'], ip_assignment, checklist_url, support_sharepoint_url):
                    print(f"  Updated ECS project notes")
                else:
                    print(f"  WARNING: Failed to update ECS notes")

                # 6. Update Engineering SOW tab from Pricing SOW
                self.update_customer_sow(project['customer'])

                # 7. Create project ticket for lifecycle tracking
                assignee = project['fields'].get('AssignedTo', 'Unassigned')
                ticket_id = self.create_project_ticket(
                    project_id=project['id'],
                    customer_name=project['customer'],
                    project_name=project['name'],
                    assignee=assignee
                )
                if ticket_id:
                    print(f"  Created project ticket: ID {ticket_id}")
                    # Link ticket back to project
                    self.update_project_ticket_id(project['id'], ticket_id)
                    print(f"  Linked ticket to project")
                else:
                    print(f"  WARNING: Failed to create project ticket")

                # 8. Create task for Kanban board
                due_date = project['fields'].get('DueDate')
                task_id = self.create_project_task(
                    project_id=project['id'],
                    customer_name=project['customer'],
                    assignee=assignee,
                    due_date=due_date
                )
                if task_id:
                    print(f"  Created Kanban task: ID {task_id}")
                    # Link task to ticket
                    if ticket_id:
                        if self.link_task_to_ticket(task_id, ticket_id):
                            print(f"  Linked task to ticket")
                else:
                    print(f"  WARNING: Failed to create Kanban task")

                # 9. Update project status to In Progress
                if self.update_project_status(project['id'], 'In Progress', 'Implementation'):
                    print(f"  Set project status: In Progress")
                else:
                    print(f"  WARNING: Failed to update project status")

                # Mark as processed
                processed_ids.add(str(project['id']))

                print(f"  DONE")
                print()

            except Exception as e:
                print(f"  ERROR: {e}")
                print()

        # Save processed list
        processed['processed'] = list(processed_ids)
        processed['last_run'] = datetime.now().isoformat()
        self.save_processed_projects(processed)

        print("=" * 60)
        print(f"Completed: {datetime.now().isoformat()}")
        print(f"Processed {len(new_installs)} projects")
        print("=" * 60)


def main():
    automation = ProjectAutomation()
    automation.process_new_projects()


if __name__ == "__main__":
    main()
