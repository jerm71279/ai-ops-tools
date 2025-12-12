#!/usr/bin/env python3
"""
New Site Install Automation
============================
Modular, repeatable process for new customer site installations.

Workflow:
1. Network discovery/scan (manual or via NetworkScannerSuite)
2. Generate scan report (manual or via existing scripts)
3. Create checklist (via network_checklist.py)
4. Generate Engineering SOW summary from Pricing SOW
5. Populate EngineeringSOW tab in customer's SOW xlsx
6. Create project + tasks in Engineering Command Center (ECC)

Usage:
    python new_site_install_automation.py --customer "Spanish Fort Water System"
    python new_site_install_automation.py --customer "SFWB" --generate-sow
    python new_site_install_automation.py --customer "SFWB" --create-project
    python new_site_install_automation.py --customer "SFWB" --all
"""
import os
import re
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import requests

# Try to import openpyxl
try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    openpyxl = None

from dotenv import load_dotenv


def parse_scan_report(scan_report_path: Path) -> Dict:
    """Parse network scan report markdown file and extract all available data"""
    data = {
        'locations': [],
        'network_range': '',
        'gateway': '',
        'gateway_manufacturer': '',
        'total_hosts': 0,
        'devices': [],
        'workstations': [],
        'printers': [],
        'switches': [],
        'access_points': [],
        'voip_devices': [],
        'servers': [],
        'security_devices': [],
        'subnet_mask': '255.255.255.0',
        'dns_primary': '',
        'has_voip': False,
        'has_security_system': False,
        'current_router': '',
        'current_switches': '',
        'current_aps': '',
    }

    if not scan_report_path.exists():
        return data

    content = scan_report_path.read_text()

    # Extract network range
    range_match = re.search(r'\*\*Network Range\*\*\s*\|\s*([^\|]+)', content)
    if range_match:
        data['network_range'] = range_match.group(1).strip()

    # Extract gateway info
    gateway_match = re.search(r'\*\*Gateway/Router\*\*\s*\|\s*([^\|]+)', content)
    if gateway_match:
        gateway_info = gateway_match.group(1).strip()
        data['gateway'] = gateway_info
        # Parse IP from gateway info
        ip_match = re.search(r'(\d+\.\d+\.\d+\.\d+)', gateway_info)
        if ip_match:
            data['dns_primary'] = ip_match.group(1)
        # Parse manufacturer
        if 'ASUS' in gateway_info:
            data['gateway_manufacturer'] = 'ASUS'
            data['current_router'] = 'ASUS Router'
        elif 'SonicWall' in gateway_info:
            data['gateway_manufacturer'] = 'SonicWall'
            data['current_router'] = 'SonicWall Firewall'

    # Extract total hosts
    hosts_match = re.search(r'\*\*Total Hosts Discovered\*\*\s*\|\s*(\d+)', content)
    if hosts_match:
        data['total_hosts'] = int(hosts_match.group(1))

    # Extract subnet mask from printer info or elsewhere
    subnet_match = re.search(r'\*\*Subnet Mask\*\*\s*\|\s*([^\|]+)', content)
    if subnet_match:
        data['subnet_mask'] = subnet_match.group(1).strip()

    # Parse device table - look for the Discovered Hosts Summary table
    device_table_pattern = r'\| IP Address \| Hostname \| Device Type \| Manufacturer \|[\s\S]*?(?=\n---|\n##|\Z)'
    device_tables = re.findall(device_table_pattern, content)

    for table in device_tables:
        rows = re.findall(r'\| ([\d\.]+) \| ([^\|]*) \| ([^\|]*) \| ([^\|]*) \|', table)
        for row in rows:
            ip, hostname, device_type, manufacturer = [x.strip() for x in row]
            device = {
                'ip': ip,
                'hostname': hostname,
                'type': device_type,
                'manufacturer': manufacturer
            }
            data['devices'].append(device)

            # Categorize devices
            type_lower = device_type.lower()
            mfr_lower = manufacturer.lower()

            if 'windows' in type_lower or 'pc' in type_lower or 'workstation' in type_lower:
                data['workstations'].append(device)
            elif 'printer' in type_lower:
                data['printers'].append(device)
            elif 'switch' in type_lower:
                data['switches'].append(device)
                if not data['current_switches']:
                    data['current_switches'] = f"{manufacturer} Switch"
                else:
                    data['current_switches'] += f", {manufacturer} Switch"
            elif 'access point' in type_lower or 'ap' in type_lower:
                data['access_points'].append(device)
                if not data['current_aps']:
                    data['current_aps'] = f"{manufacturer} AP"
            elif 'voip' in type_lower or 'phone' in type_lower or 'yealink' in mfr_lower:
                data['voip_devices'].append(device)
                data['has_voip'] = True
            elif 'server' in type_lower:
                data['servers'].append(device)
            elif 'security' in type_lower or 'alarm' in type_lower or 'panel' in type_lower or 'digital monitoring' in mfr_lower:
                data['security_devices'].append(device)
                data['has_security_system'] = True

    # Count locations from section headers
    location_matches = re.findall(r'### Location #\d+ - ([^\n]+)', content)
    data['locations'] = location_matches

    return data


def parse_sow_for_customer_info(sow_path: Path) -> Dict:
    """Extract customer contact info from SOW file"""
    info = {
        'customer_name': '',
        'address': '',
        'city_state_zip': '',
        'phone': '',
        'email': '',
        'contact_name': '',
        'project_name': '',
    }

    if not sow_path or not sow_path.exists() or not openpyxl:
        return info

    try:
        wb = openpyxl.load_workbook(sow_path)

        # Check SOW sheet for customer info
        if 'SOW' in wb.sheetnames:
            ws = wb['SOW']
            for row in ws.iter_rows(min_row=1, max_row=20, max_col=6, values_only=True):
                if row[1] and isinstance(row[1], str):
                    label = row[1].strip().lower()
                    # Customer info is typically in column B (label) with value in merged cells
                    if label == 'name' and row[2]:
                        info['customer_name'] = str(row[2]).strip()
                    elif label == 'address' and row[2]:
                        info['address'] = str(row[2]).strip()
                    elif 'city' in label and row[2]:
                        info['city_state_zip'] = str(row[2]).strip()
                    elif label == 'phone' and row[2]:
                        info['phone'] = str(row[2]).strip()
                    elif label == 'email' and row[2]:
                        info['email'] = str(row[2]).strip()
                    elif 'project' in label and row[2]:
                        info['project_name'] = str(row[2]).strip()

        wb.close()
    except Exception as e:
        print(f"Warning: Could not parse SOW for customer info: {e}")

    return info


def parse_sow_for_equipment(sow_path: Path) -> Dict:
    """Extract recommended equipment from Pricing SOW"""
    equipment = {
        'recommended_router': '',
        'recommended_switches': [],
        'recommended_aps': [],
        'locations': {},
    }

    if not sow_path or not sow_path.exists() or not openpyxl:
        return equipment

    try:
        wb = openpyxl.load_workbook(sow_path)

        if 'Pricing SOW' in wb.sheetnames:
            ws = wb['Pricing SOW']
            current_location = 'Main'

            for row in ws.iter_rows(min_row=2, max_row=50, max_col=6, values_only=True):
                desc = str(row[0]).strip() if row[0] else ''
                mfr = str(row[1]).strip() if row[1] else ''
                model = str(row[2]).strip() if row[2] else ''
                qty = row[4] if row[4] else 1

                if not desc or desc == 'None' or desc.startswith('='):
                    continue

                # Check for location header
                if any(kw in desc.lower() for kw in ['office', 'road', 'location', 'site']) and not mfr:
                    current_location = desc
                    equipment['locations'][current_location] = []
                    continue

                desc_lower = desc.lower()

                # Identify equipment types
                if 'firewall' in desc_lower or 'sonicwall' in mfr.lower():
                    equipment['recommended_router'] = f"{mfr} {model}".strip()
                elif 'switch' in desc_lower:
                    equipment['recommended_switches'].append(f"{mfr} {model}".strip())
                elif 'ap' in desc_lower or 'u7' in model.lower() or 'access point' in desc_lower:
                    try:
                        qty_int = int(qty)
                    except:
                        qty_int = 1
                    equipment['recommended_aps'].append(f"{qty_int}x {mfr} {model}".strip())

        wb.close()
    except Exception as e:
        print(f"Warning: Could not parse SOW for equipment: {e}")

    return equipment

load_dotenv('/home/mavrick/Projects/Secondbrain/.env')

# Paths
SHAREPOINT_LOCAL = Path("/home/mavrick/Projects/Secondbrain/input_documents/sharepoint_all")
OBERACONNECT_TECH = SHAREPOINT_LOCAL / "oberaconnect_technical/Documents"
OUTPUT_DIR = Path("/home/mavrick/Projects/Secondbrain/output")
SCANS_DIR = Path("/home/mavrick/Projects/NetworkScannerSuite/service/scans")
SUBNET_ALLOCATOR_PATH = Path("/home/mavrick/Projects/network-config-builder/subnet_allocator.py")
SUBNET_REGISTRY_PATH = Path("/home/mavrick/Projects/network-config-builder/CUSTOMER_SUBNET_TRACKER.csv")

# Import subnet allocator if available
sys.path.insert(0, str(SUBNET_ALLOCATOR_PATH.parent))
try:
    from subnet_allocator import SubnetAllocator
    SUBNET_ALLOCATOR_AVAILABLE = True
except ImportError:
    SUBNET_ALLOCATOR_AVAILABLE = False
    SubnetAllocator = None

# Customer aliases
CUSTOMER_ALIASES = {
    'SFWB': 'Spanish Fort Water System',
    'SFWS': 'Spanish Fort Water System',
    'Spanish Fort Waterboard': 'Spanish Fort Water System',
}


class SharePointClient:
    """Client for SharePoint Graph API operations"""

    def __init__(self):
        self.tenant_id = os.getenv('AZURE_TENANT_ID')
        self.client_id = os.getenv('AZURE_CLIENT_ID')
        self.client_secret = os.getenv('AZURE_CLIENT_SECRET')
        self.base_url = "https://graph.microsoft.com/v1.0"
        self.token = None
        self.site_id = None
        self.drive_id = None

    def authenticate(self):
        """Get access token"""
        url = f"https://login.microsoftonline.com/{self.tenant_id}/oauth2/v2.0/token"
        data = {
            'grant_type': 'client_credentials',
            'client_id': self.client_id,
            'client_secret': self.client_secret,
            'scope': 'https://graph.microsoft.com/.default'
        }
        response = requests.post(url, data=data)
        if response.status_code == 200:
            self.token = response.json().get('access_token')
            return True
        return False

    def get_site_and_drive(self, site_name="OberaConnect Technical"):
        """Get site and drive IDs"""
        if not self.token:
            self.authenticate()

        headers = {'Authorization': f'Bearer {self.token}'}
        resp = requests.get(f"{self.base_url}/sites?search=*", headers=headers)
        sites = resp.json().get('value', [])

        for site in sites:
            if site.get('displayName') == site_name:
                self.site_id = site['id']
                resp = requests.get(f"{self.base_url}/sites/{self.site_id}/drives", headers=headers)
                drives = resp.json().get('value', [])
                for drive in drives:
                    if drive['name'] == 'Documents':
                        self.drive_id = drive['id']
                        return True
        return False

    def upload_file(self, folder_path, filename, content):
        """Upload file to SharePoint"""
        if not self.token:
            self.authenticate()
        if not self.site_id:
            self.get_site_and_drive()

        headers = {
            'Authorization': f'Bearer {self.token}',
            'Content-Type': 'application/octet-stream'
        }
        url = f"{self.base_url}/sites/{self.site_id}/drives/{self.drive_id}/root:/{folder_path}/{filename}:/content"
        response = requests.put(url, headers=headers, data=content)
        return response.status_code in [200, 201], response.json() if response.ok else None


def resolve_customer_name(name):
    """Resolve customer alias to full name"""
    return CUSTOMER_ALIASES.get(name.upper(), CUSTOMER_ALIASES.get(name, name))


def get_customer_subnet_allocation(customer_name: str) -> Optional[Dict]:
    """Look up existing subnet allocation for a customer from the registry"""
    if not SUBNET_REGISTRY_PATH.exists():
        return None

    import csv
    with open(SUBNET_REGISTRY_PATH, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get('Customer Name', '').lower() == customer_name.lower():
                return {
                    'assigned_subnet': row.get('Configured LAN Subnet', ''),
                    'assigned_gateway': row.get('LAN Gateway', ''),
                    'assigned_subnet_mask': '255.255.255.0',  # /24
                    'dhcp_range': row.get('DHCP Range', ''),
                    'wan_ip': row.get('WAN IP', ''),
                    'wan_gateway': row.get('WAN Gateway', ''),
                    'allocated_block': row.get('Allocated Subnet Block', ''),
                }
    return None


def allocate_new_subnet(customer_name: str, customer_id: str = None,
                        wan_ip: str = '', wan_gateway: str = '',
                        location: str = '', circuit_id: str = '') -> Optional[Dict]:
    """Allocate a new subnet block for a customer"""
    if not SUBNET_ALLOCATOR_AVAILABLE:
        print("Warning: Subnet allocator not available")
        return None

    allocator = SubnetAllocator(str(SUBNET_REGISTRY_PATH))

    # Generate customer ID if not provided
    if not customer_id:
        # Use first letters of each word + number
        words = customer_name.split()
        customer_id = ''.join(w[0].upper() for w in words if w)
        # Add a number based on existing allocations
        existing = allocator.get_allocated_blocks()
        customer_id = f"{customer_id}{len(existing) + 1:03d}"

    try:
        allocation = allocator.allocate_block(
            customer_id=customer_id,
            customer_name=customer_name,
            wan_ip=wan_ip or '[TBD]',
            wan_gateway=wan_gateway or '[TBD]',
            location=location or customer_name,
            circuit_id=circuit_id or '[TBD]',
            notes=f"Allocated via automation for {customer_name}"
        )

        return {
            'assigned_subnet': allocation['configured_subnet'],
            'assigned_gateway': allocation['gateway'],
            'assigned_subnet_mask': '255.255.255.0',
            'dhcp_range': allocation['dhcp_range'],
            'wan_ip': allocation['wan_ip'],
            'wan_gateway': allocation['wan_gateway'],
            'allocated_block': allocation['allocated_block'],
            'customer_id': customer_id,
        }
    except ValueError as e:
        print(f"Error allocating subnet: {e}")
        return None


def get_next_available_subnet() -> Optional[Dict]:
    """Get the next available subnet block without allocating it"""
    if not SUBNET_ALLOCATOR_AVAILABLE:
        return None

    try:
        allocator = SubnetAllocator(str(SUBNET_REGISTRY_PATH))
        block_range, subnet, gateway, dhcp_start, dhcp_end = allocator.get_next_available_block()
        return {
            'block_range': block_range,
            'subnet': subnet,
            'gateway': gateway,
            'dhcp_range': f"{dhcp_start}-{dhcp_end.split('.')[-1]}",
        }
    except ValueError:
        return None


def find_scan_files(customer_name):
    """Find scan files for a customer in various locations"""
    customer_name = resolve_customer_name(customer_name)
    scan_files = []

    # Search patterns for customer name variations
    search_terms = [
        customer_name.lower().replace(' ', '_'),
        customer_name.lower().replace(' ', '-'),
        customer_name.lower().replace(' ', ''),
    ]
    # Add alias variations
    for alias, full_name in CUSTOMER_ALIASES.items():
        if full_name == customer_name:
            search_terms.append(alias.lower())

    # Check NetworkScannerSuite scans directory
    if SCANS_DIR.exists():
        for scan_folder in SCANS_DIR.iterdir():
            if scan_folder.is_dir():
                folder_lower = scan_folder.name.lower()
                if any(term in folder_lower for term in search_terms):
                    # Get all scan files from this folder
                    for f in scan_folder.glob("*"):
                        if f.is_file() and f.suffix in ['.nmap', '.xml', '.gnmap', '.html', '.md', '.txt']:
                            scan_files.append(f)

    # Check output directory for reports
    if OUTPUT_DIR.exists():
        for f in OUTPUT_DIR.glob("*"):
            if f.is_file():
                f_lower = f.name.lower()
                if any(term in f_lower for term in search_terms):
                    if 'scan' in f_lower or 'checklist' in f_lower:
                        scan_files.append(f)

    return scan_files


def find_scan_report(customer_name) -> Optional[Path]:
    """Find the main scan report markdown file for a customer"""
    customer_name = resolve_customer_name(customer_name)

    # Search patterns
    search_terms = [
        customer_name.lower().replace(' ', '_'),
        customer_name.lower().replace(' ', '-'),
        customer_name.lower().replace(' ', ''),
    ]
    for alias, full_name in CUSTOMER_ALIASES.items():
        if full_name == customer_name:
            search_terms.append(alias.lower())

    # Check output directory for scan reports
    if OUTPUT_DIR.exists():
        for f in OUTPUT_DIR.glob("*Scan*Report*.md"):
            f_lower = f.name.lower()
            if any(term in f_lower for term in search_terms):
                return f

    # Check customer SharePoint folder
    customer_folder = find_customer_folder(customer_name)
    if customer_folder:
        tech_docs = customer_folder / "Technical Docs"
        if tech_docs.exists():
            for f in tech_docs.glob("*Scan*Report*.md"):
                return f
            for f in tech_docs.glob("*scan*.md"):
                return f

    return None


def gather_checklist_data(customer_name) -> Dict:
    """Gather all available data for checklist auto-fill from multiple sources"""
    customer_name = resolve_customer_name(customer_name)

    data = {
        # Project Info
        'customer_name': customer_name,
        'date': datetime.now().strftime('%Y-%m-%d'),
        'project_manager': '',
        'site_address': '',
        'contact_name': '',
        'contact_email': '',
        'contact_phone': '',

        # Current Network (from scan)
        'current_isp': '',
        'internet_speed': '',
        'num_users': '',
        'num_devices': '',
        'current_router': '',
        'current_switches': '',
        'current_aps': '',

        # Current/Existing Network Configuration (from scan)
        'existing_subnet': '',
        'existing_gateway': '',
        'existing_subnet_mask': '',
        'existing_dns': '',

        # New OberaConnect Assigned Network Configuration
        'assigned_subnet': '',
        'assigned_gateway': '',
        'assigned_subnet_mask': '',
        'assigned_dns_primary': '',
        'assigned_dns_secondary': '',

        # Public IP (from ISP)
        'public_ip': '',
        'public_ip_range': '',

        # Requirements
        'vlans_needed': 'Data, Voice, Guest, IoT, Security',
        'guest_network': 'Yes',
        'voip_required': 'No',
        'vpn_required': 'Yes',

        # Recommended Equipment (from SOW)
        'recommended_router': 'SonicWall TZ',
        'recommended_switches': '',
        'recommended_aps': 'Ubiquiti UniFi',

        # Device counts (from scan)
        'workstation_count': 0,
        'printer_count': 0,
        'voip_count': 0,
        'server_count': 0,
        'locations': [],

        # Notes
        'notes': '',
        'auto_filled_fields': [],
    }

    print(f"\n--- Gathering data for {customer_name} ---")

    # 1. Parse scan report
    scan_report = find_scan_report(customer_name)
    if scan_report:
        print(f"  Found scan report: {scan_report.name}")
        scan_data = parse_scan_report(scan_report)

        if scan_data['network_range']:
            data['existing_subnet'] = scan_data['network_range']
            data['auto_filled_fields'].append('existing_subnet')

        if scan_data['gateway']:
            # Extract IP for gateway
            ip_match = re.search(r'(\d+\.\d+\.\d+\.\d+)', scan_data['gateway'])
            if ip_match:
                data['existing_gateway'] = ip_match.group(1)
                data['existing_dns'] = ip_match.group(1)  # Often same as gateway on existing networks
                data['auto_filled_fields'].extend(['existing_gateway', 'existing_dns'])

        if scan_data['subnet_mask']:
            data['existing_subnet_mask'] = scan_data['subnet_mask']
            data['auto_filled_fields'].append('existing_subnet_mask')

        if scan_data['current_router']:
            data['current_router'] = scan_data['current_router']
            data['auto_filled_fields'].append('current_router')

        if scan_data['current_switches']:
            data['current_switches'] = scan_data['current_switches']
            data['auto_filled_fields'].append('current_switches')

        if scan_data['current_aps']:
            data['current_aps'] = scan_data['current_aps']
            data['auto_filled_fields'].append('current_aps')

        # Device counts
        data['workstation_count'] = len(scan_data['workstations'])
        data['printer_count'] = len(scan_data['printers'])
        data['voip_count'] = len(scan_data['voip_devices'])
        data['server_count'] = len(scan_data['servers'])

        data['num_devices'] = str(scan_data['total_hosts'])
        data['auto_filled_fields'].append('num_devices')

        # Estimate users from workstations
        if scan_data['workstations']:
            data['num_users'] = str(len(scan_data['workstations']))
            data['auto_filled_fields'].append('num_users')

        # VoIP detection
        if scan_data['has_voip']:
            data['voip_required'] = 'Yes'
            data['auto_filled_fields'].append('voip_required')

        # Locations
        if scan_data['locations']:
            data['locations'] = scan_data['locations']

        print(f"    - Existing Network: {data['existing_subnet']}")
        print(f"    - Existing Gateway: {data['existing_gateway']}")
        print(f"    - Devices: {data['num_devices']} ({data['workstation_count']} PCs, {data['printer_count']} printers)")
        print(f"    - VoIP: {data['voip_required']}")
    else:
        print("  No scan report found")

    # 2. Parse SOW for customer info and equipment
    customer_folder = find_customer_folder(customer_name)
    sow_file = find_sow_file(customer_folder) if customer_folder else None

    if sow_file:
        print(f"  Found SOW: {sow_file.name}")

        # Customer info
        customer_info = parse_sow_for_customer_info(sow_file)
        if customer_info['address']:
            data['site_address'] = customer_info['address']
            if customer_info['city_state_zip']:
                data['site_address'] += f", {customer_info['city_state_zip']}"
            data['auto_filled_fields'].append('site_address')
        if customer_info['phone']:
            data['contact_phone'] = customer_info['phone']
            data['auto_filled_fields'].append('contact_phone')
        if customer_info['email']:
            data['contact_email'] = customer_info['email']
            data['auto_filled_fields'].append('contact_email')

        # Equipment recommendations
        equipment = parse_sow_for_equipment(sow_file)
        if equipment['recommended_router']:
            data['recommended_router'] = equipment['recommended_router']
            data['auto_filled_fields'].append('recommended_router')
            print(f"    - Recommended Router: {data['recommended_router']}")

        if equipment['recommended_switches']:
            data['recommended_switches'] = ', '.join(equipment['recommended_switches'])
            data['auto_filled_fields'].append('recommended_switches')
            print(f"    - Recommended Switches: {data['recommended_switches']}")

        if equipment['recommended_aps']:
            data['recommended_aps'] = ', '.join(equipment['recommended_aps'])
            data['auto_filled_fields'].append('recommended_aps')
            print(f"    - Recommended APs: {data['recommended_aps']}")
    else:
        print("  No SOW file found")

    # 3. Check for existing subnet allocation
    subnet_alloc = get_customer_subnet_allocation(customer_name)
    if subnet_alloc:
        print(f"  Found existing subnet allocation")
        data['assigned_subnet'] = subnet_alloc['assigned_subnet']
        data['assigned_gateway'] = subnet_alloc['assigned_gateway']
        data['assigned_subnet_mask'] = subnet_alloc['assigned_subnet_mask']
        data['auto_filled_fields'].extend(['assigned_subnet', 'assigned_gateway', 'assigned_subnet_mask'])
        print(f"    - Assigned Subnet: {data['assigned_subnet']}")
        print(f"    - Assigned Gateway: {data['assigned_gateway']}")

        if subnet_alloc.get('dhcp_range'):
            data['dhcp_range'] = subnet_alloc['dhcp_range']
    else:
        # Show what the next available subnet would be
        next_subnet = get_next_available_subnet()
        if next_subnet:
            print(f"  No subnet allocated yet. Next available: {next_subnet['subnet']}")
        else:
            print("  No subnet allocated (subnet allocator not available)")

    # Summary
    print(f"\n  Auto-filled {len(data['auto_filled_fields'])} fields: {', '.join(data['auto_filled_fields'])}")

    return data


def generate_checklist_html(data: Dict) -> str:
    """Generate HTML checklist with auto-filled customer data"""

    # Read the template checklist
    template_path = Path("/home/mavrick/Projects/network_install_workflow_checklist.md")
    if not template_path.exists():
        print(f"Warning: Template not found at {template_path}")
        template_content = "## Checklist\n- [ ] No template found"
    else:
        template_content = template_path.read_text()

    # Convert markdown checklist to HTML
    checklist_html = ""
    for line in template_content.split('\n'):
        if line.startswith('# '):
            continue  # Skip main title
        elif line.startswith('## '):
            checklist_html += f'<h2>{line[3:]}</h2>\n'
        elif line.startswith('### '):
            checklist_html += f'<h3>{line[4:]}</h3>\n'
        elif line.startswith('- [ ] '):
            text = line[6:]
            checklist_html += f'<div class="checklist-item"><label><input type="checkbox"> {text}</label></div>\n'
        elif line.startswith('  - [ ] '):
            text = line[8:]
            checklist_html += f'<div class="checklist-item nested"><label><input type="checkbox"> {text}</label></div>\n'
        elif line.startswith('---'):
            checklist_html += '<hr>\n'
        elif line.strip():
            checklist_html += f'<p>{line}</p>\n'

    # Build auto-filled indicator
    auto_filled = data.get('auto_filled_fields', [])

    def field_class(field_name):
        return 'auto-filled' if field_name in auto_filled else ''

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Network Checklist - {data['customer_name']}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
            background: #f5f5f5;
        }}
        .header {{
            padding: 20px 0;
            margin-bottom: 20px;
            border-bottom: 3px solid #4CAF50;
        }}
        .header h1 {{
            margin: 10px 0;
            font-size: 24px;
            color: #1e3a5f;
        }}
        .auto-filled {{
            background: #e8f5e9;
            border-left: 3px solid #4CAF50;
            padding-left: 8px;
        }}
        .manual-entry {{
            background: #fff3e0;
            border-left: 3px solid #ff9800;
            padding-left: 8px;
        }}
        .customer-info {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .customer-info h2 {{
            margin-top: 0;
            color: #1e3a5f;
            border-bottom: 2px solid #4CAF50;
            padding-bottom: 10px;
        }}
        .info-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 15px;
        }}
        .info-item {{
            padding: 8px;
            border-radius: 4px;
        }}
        .info-item strong {{
            color: #1e3a5f;
        }}
        .legend {{
            display: flex;
            gap: 20px;
            margin-bottom: 15px;
            font-size: 12px;
        }}
        .legend-item {{
            display: flex;
            align-items: center;
            gap: 5px;
        }}
        .legend-auto {{
            width: 12px;
            height: 12px;
            background: #e8f5e9;
            border-left: 3px solid #4CAF50;
        }}
        .legend-manual {{
            width: 12px;
            height: 12px;
            background: #fff3e0;
            border-left: 3px solid #ff9800;
        }}
        .checklist-section {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h2 {{
            color: #1e3a5f;
            border-bottom: 2px solid #4CAF50;
            padding-bottom: 10px;
            margin-top: 30px;
        }}
        h3 {{
            color: #2d5a87;
            margin-top: 20px;
        }}
        .checklist-item {{
            padding: 8px 0;
            border-bottom: 1px solid #eee;
        }}
        .checklist-item:last-child {{
            border-bottom: none;
        }}
        .checklist-item.nested {{
            padding-left: 30px;
        }}
        .checklist-item label {{
            cursor: pointer;
            display: flex;
            align-items: flex-start;
        }}
        .checklist-item input[type="checkbox"] {{
            margin-right: 10px;
            margin-top: 3px;
            width: 18px;
            height: 18px;
            cursor: pointer;
        }}
        hr {{
            border: none;
            border-top: 2px solid #4CAF50;
            margin: 30px 0;
        }}
        .progress-bar {{
            background: #e0e0e0;
            border-radius: 10px;
            height: 20px;
            margin: 20px 0;
            overflow: hidden;
        }}
        .progress-fill {{
            background: linear-gradient(90deg, #4CAF50, #45a049);
            height: 100%;
            width: 0%;
            transition: width 0.3s ease;
            border-radius: 10px;
        }}
        .progress-text {{
            text-align: center;
            font-weight: bold;
            color: #1e3a5f;
        }}
        .device-summary {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 10px;
            margin-top: 15px;
        }}
        .device-count {{
            text-align: center;
            padding: 10px;
            background: #f5f5f5;
            border-radius: 8px;
        }}
        .device-count .number {{
            font-size: 24px;
            font-weight: bold;
            color: #1e3a5f;
        }}
        .device-count .label {{
            font-size: 12px;
            color: #666;
        }}
        @media print {{
            body {{ background: white; }}
            .checklist-section, .customer-info {{ box-shadow: none; break-inside: avoid; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Network Installation Checklist</h1>
        <div>Customer: {data['customer_name']}</div>
        <div>Date: {data['date']}</div>
        <div class="legend">
            <div class="legend-item"><div class="legend-auto"></div> Auto-filled from scan/SOW</div>
            <div class="legend-item"><div class="legend-manual"></div> Requires manual entry</div>
        </div>
    </div>

    <div class="customer-info">
        <h2>Project Overview</h2>
        <div class="info-grid">
            <div class="info-item {field_class('project_manager')} {'manual-entry' if 'project_manager' not in auto_filled else ''}"><strong>Project Manager:</strong> {data['project_manager'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('site_address')} {'manual-entry' if 'site_address' not in auto_filled else ''}"><strong>Site Address:</strong> {data['site_address'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('contact_name')} {'manual-entry' if 'contact_name' not in auto_filled else ''}"><strong>Contact:</strong> {data['contact_name'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('contact_email')} {'manual-entry' if 'contact_email' not in auto_filled else ''}"><strong>Email:</strong> {data['contact_email'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('contact_phone')} {'manual-entry' if 'contact_phone' not in auto_filled else ''}"><strong>Phone:</strong> {data['contact_phone'] or '[MANUAL ENTRY]'}</div>
        </div>

        <h3>Current Network</h3>
        <div class="info-grid">
            <div class="info-item {'manual-entry' if 'current_isp' not in auto_filled else field_class('current_isp')}"><strong>ISP:</strong> {data['current_isp'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {'manual-entry' if 'internet_speed' not in auto_filled else field_class('internet_speed')}"><strong>Speed:</strong> {data['internet_speed'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('num_users')} {'manual-entry' if 'num_users' not in auto_filled else ''}"><strong>Users:</strong> {data['num_users'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('num_devices')}"><strong>Devices:</strong> {data['num_devices'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('current_router')}"><strong>Router:</strong> {data['current_router'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('current_switches')}"><strong>Switches:</strong> {data['current_switches'] or 'None detected'}</div>
            <div class="info-item {field_class('current_aps')}"><strong>APs:</strong> {data['current_aps'] or 'None detected'}</div>
        </div>

        <div class="device-summary">
            <div class="device-count"><div class="number">{data['workstation_count']}</div><div class="label">Workstations</div></div>
            <div class="device-count"><div class="number">{data['printer_count']}</div><div class="label">Printers</div></div>
            <div class="device-count"><div class="number">{data['voip_count']}</div><div class="label">VoIP Devices</div></div>
            <div class="device-count"><div class="number">{data['server_count']}</div><div class="label">Servers</div></div>
        </div>

        <h3>Existing Network Configuration (from scan)</h3>
        <div class="info-grid">
            <div class="info-item {field_class('existing_subnet')}"><strong>Existing Subnet:</strong> {data['existing_subnet'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('existing_gateway')}"><strong>Existing Gateway:</strong> {data['existing_gateway'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('existing_subnet_mask')}"><strong>Existing Subnet Mask:</strong> {data['existing_subnet_mask'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {field_class('existing_dns')}"><strong>Existing DNS:</strong> {data['existing_dns'] or '[MANUAL ENTRY]'}</div>
        </div>

        <h3>OberaConnect Assigned Network Configuration</h3>
        <div class="info-grid">
            <div class="info-item {'manual-entry' if 'assigned_subnet' not in auto_filled else field_class('assigned_subnet')}"><strong>Assigned Subnet:</strong> {data['assigned_subnet'] or '[FROM SUBNET GENERATOR]'}</div>
            <div class="info-item {'manual-entry' if 'assigned_gateway' not in auto_filled else field_class('assigned_gateway')}"><strong>Assigned Gateway:</strong> {data['assigned_gateway'] or '[FROM SUBNET GENERATOR]'}</div>
            <div class="info-item {'manual-entry' if 'assigned_subnet_mask' not in auto_filled else field_class('assigned_subnet_mask')}"><strong>Assigned Subnet Mask:</strong> {data['assigned_subnet_mask'] or '[FROM SUBNET GENERATOR]'}</div>
            <div class="info-item {'manual-entry' if 'assigned_dns_primary' not in auto_filled else field_class('assigned_dns_primary')}"><strong>Primary DNS:</strong> {data['assigned_dns_primary'] or '[FROM SUBNET GENERATOR]'}</div>
            <div class="info-item {'manual-entry' if 'assigned_dns_secondary' not in auto_filled else field_class('assigned_dns_secondary')}"><strong>Secondary DNS:</strong> {data['assigned_dns_secondary'] or '[FROM SUBNET GENERATOR]'}</div>
        </div>

        <h3>Public IP (from ISP)</h3>
        <div class="info-grid">
            <div class="info-item {'manual-entry' if 'public_ip' not in auto_filled else field_class('public_ip')}"><strong>Public IP:</strong> {data['public_ip'] or '[MANUAL ENTRY]'}</div>
            <div class="info-item {'manual-entry' if 'public_ip_range' not in auto_filled else field_class('public_ip_range')}"><strong>Public IP Range:</strong> {data['public_ip_range'] or '[MANUAL ENTRY]'}</div>
        </div>

        <h3>Requirements</h3>
        <div class="info-grid">
            <div class="info-item"><strong>VLANs:</strong> {data['vlans_needed']}</div>
            <div class="info-item"><strong>Guest Network:</strong> {data['guest_network']}</div>
            <div class="info-item {field_class('voip_required')}"><strong>VoIP:</strong> {data['voip_required']}</div>
            <div class="info-item"><strong>VPN:</strong> {data['vpn_required']}</div>
        </div>

        <h3>Recommended Equipment (from SOW)</h3>
        <div class="info-grid">
            <div class="info-item {field_class('recommended_router')}"><strong>Router/Firewall:</strong> {data['recommended_router']}</div>
            <div class="info-item {field_class('recommended_switches')}"><strong>Switches:</strong> {data['recommended_switches'] or '[TBD]'}</div>
            <div class="info-item {field_class('recommended_aps')}"><strong>Access Points:</strong> {data['recommended_aps']}</div>
        </div>
    </div>

    <div class="progress-text">Progress: <span id="progress-percent">0%</span></div>
    <div class="progress-bar">
        <div class="progress-fill" id="progress-fill"></div>
    </div>

    <div class="checklist-section">
        {checklist_html}
    </div>

    <script>
        function updateProgress() {{
            const checkboxes = document.querySelectorAll('input[type="checkbox"]');
            const checked = document.querySelectorAll('input[type="checkbox"]:checked');
            const percent = Math.round((checked.length / checkboxes.length) * 100);
            document.getElementById('progress-percent').textContent = percent + '%';
            document.getElementById('progress-fill').style.width = percent + '%';
        }}
        document.querySelectorAll('input[type="checkbox"]').forEach(cb => {{
            cb.addEventListener('change', updateProgress);
        }});
        updateProgress();
    </script>
</body>
</html>'''

    return html


def upload_scans_to_sharepoint(customer_name, scan_files, sp_client, dry_run=False):
    """Upload scan files to SharePoint Technical Docs/Scans folder"""
    customer_name = resolve_customer_name(customer_name)
    folder_path = f"{customer_name}/Technical Docs/Scans"

    uploaded = []
    failed = []

    for scan_file in scan_files:
        if dry_run:
            print(f"  [DRY RUN] Would upload: {scan_file.name}")
            uploaded.append(scan_file.name)
            continue

        try:
            with open(scan_file, 'rb') as f:
                content = f.read()
            success, result = sp_client.upload_file(folder_path, scan_file.name, content)
            if success:
                print(f"  Uploaded: {scan_file.name}")
                uploaded.append(scan_file.name)
            else:
                print(f"  Failed: {scan_file.name}")
                failed.append(scan_file.name)
        except Exception as e:
            print(f"  Error uploading {scan_file.name}: {e}")
            failed.append(scan_file.name)

    return uploaded, failed


def find_customer_folder(customer_name):
    """Find customer folder in local SharePoint sync"""
    customer_name = resolve_customer_name(customer_name)

    # Check OberaConnect Technical
    for folder in OBERACONNECT_TECH.iterdir():
        if folder.is_dir() and customer_name.lower() in folder.name.lower():
            return folder
    return None


def find_sow_file(customer_folder, sow_name=None):
    """Find SOW xlsx file in customer folder

    Args:
        customer_folder: Path to customer folder
        sow_name: Optional specific SOW filename to look for (e.g., 'SOW-SFWS.xlsx')
    """
    # Check multiple possible SOW folder names
    sow_folder_names = ["Statements of Work (SOWs)", "SOWs", "SOW"]
    sow_folder = None

    for folder_name in sow_folder_names:
        potential = customer_folder / folder_name
        if potential.exists():
            sow_folder = potential
            break

    # Also check for SOW.xlsx directly in customer folder
    direct_sow = customer_folder / "SOW.xlsx"
    if direct_sow.exists():
        return direct_sow

    if sow_folder and sow_folder.exists():
        # If specific name provided, look for that first
        if sow_name:
            specific = sow_folder / sow_name
            if specific.exists():
                return specific

        # Look for main SOW file (starts with 'SOW-' not containing 'Camera')
        for f in sow_folder.glob("SOW-*.xlsx"):
            if 'camera' not in f.name.lower():
                return f

        # Look for SOW.xlsx
        for f in sow_folder.glob("SOW.xlsx"):
            return f

        # Fallback: any SOW file not containing 'Camera'
        for f in sow_folder.glob("*.xlsx"):
            if 'SOW' in f.name.upper() and 'camera' not in f.name.lower():
                return f
    return None


def extract_equipment_from_pricing(wb):
    """Extract equipment list from Pricing SOW sheet"""
    equipment = {'locations': {}}

    if 'Pricing SOW' not in wb.sheetnames:
        return equipment

    ws = wb['Pricing SOW']
    current_location = 'Main'

    for row in ws.iter_rows(min_row=2, max_row=50, max_col=5):
        desc = row[0].value
        mfr = row[1].value
        model = row[2].value
        qty = row[4].value

        if desc:
            desc_str = str(desc).strip()

            # Skip formula results, totals, and empty/placeholder rows
            if desc_str.startswith('=') or 'total' in desc_str.lower() or 'labor' in desc_str.lower() or 'e/m' in desc_str.lower():
                continue
            if desc_str == 'Description' or desc_str == 'Manufacturer':
                continue

            # Check if this is a location header
            if any(kw in desc_str.lower() for kw in ['office', 'road', 'location', 'site', 'building']):
                if not mfr and not model:
                    current_location = desc_str
                    equipment['locations'][current_location] = []
                    continue

            # Add equipment item (only if has manufacturer/model or is meaningful)
            # Skip if manufacturer or model contains formulas
            mfr_str = str(mfr).strip() if mfr else ''
            model_str = str(model).strip() if model else ''
            if mfr_str.startswith('=') or model_str.startswith('='):
                continue

            if mfr_str or model_str:
                if current_location not in equipment['locations']:
                    equipment['locations'][current_location] = []

                equipment['locations'][current_location].append({
                    'description': desc_str,
                    'manufacturer': mfr_str,
                    'model': model_str,
                    'quantity': int(qty) if qty and str(qty).isdigit() else 1
                })

    return equipment


def generate_engineering_sow_summary(equipment, customer_name, scan_data=None):
    """Generate Engineering SOW summary text"""
    lines = []
    lines.append(f"Network Infrastructure Installation for {customer_name}")
    lines.append("")

    for location, items in equipment['locations'].items():
        if not items:
            continue

        lines.append(f"{location}:")

        # Group by type
        firewalls = [i for i in items if 'firewall' in i['description'].lower() or 'sonicwall' in i['manufacturer'].lower()]
        switches = [i for i in items if 'switch' in i['description'].lower()]
        aps = [i for i in items if 'ap' in i['description'].lower() or 'u7' in i['model'].lower() or 'access point' in i['description'].lower()]
        cabling = [i for i in items if 'cat' in i['description'].lower() or 'drop' in i['description'].lower()]
        other = [i for i in items if i not in firewalls + switches + aps + cabling]

        if firewalls:
            for item in firewalls:
                lines.append(f"  - Install and configure {item['manufacturer']} {item['model']} firewall")
                lines.append(f"    - Set up firewall rules and security policies")
                lines.append(f"    - Configure VPN (if required)")

        if switches:
            for item in switches:
                qty = item['quantity']
                lines.append(f"  - Install {qty}x {item['manufacturer']} {item['model']} switch{'es' if qty > 1 else ''}")
                lines.append(f"    - Configure ports and PoE")

        if aps:
            total_aps = sum(i['quantity'] for i in aps)
            lines.append(f"  - Install {total_aps}x UniFi access point{'s' if total_aps > 1 else ''}")
            lines.append(f"    - Configure SSIDs and wireless security")
            lines.append(f"    - Optimize placement for coverage")

        if cabling:
            total_drops = sum(i['quantity'] for i in cabling)
            lines.append(f"  - Install {total_drops}x CAT6 drop{'s' if total_drops > 1 else ''}")

        for item in other:
            if item['description'] and item['description'] not in ['', ' ']:
                lines.append(f"  - {item['description']}")

        lines.append("")

    # Calculate engineering hours
    num_sites = len([loc for loc, items in equipment['locations'].items() if items])
    num_firewalls = sum(
        len([i for i in items if 'firewall' in i['description'].lower() or 'sonicwall' in i['manufacturer'].lower()])
        for items in equipment['locations'].values()
    )
    has_other_devices = any(
        any(i for i in items if 'firewall' not in i['description'].lower() and 'sonicwall' not in i['manufacturer'].lower())
        for items in equipment['locations'].values()
    )

    discovery_hours = num_sites  # 1 hour per site
    firewall_hours = num_firewalls  # 1 hour per firewall
    other_device_hours = 1 if has_other_devices else 0  # 1 hour for all other devices combined
    onsite_hours = num_sites  # 1 hour per site for verification

    total_hours = discovery_hours + firewall_hours + other_device_hours + onsite_hours

    lines.append("Engineering Hours Estimate:")
    lines.append(f"  - Network Discovery: {discovery_hours} hour{'s' if discovery_hours != 1 else ''} ({num_sites} site{'s' if num_sites != 1 else ''})")
    if num_firewalls > 0:
        lines.append(f"  - Firewall Configuration: {firewall_hours} hour{'s' if firewall_hours != 1 else ''} ({num_firewalls} firewall{'s' if num_firewalls != 1 else ''})")
    if has_other_devices:
        lines.append(f"  - Other Device Configuration: {other_device_hours} hour")
    lines.append(f"  - Onsite Install/Verification: {onsite_hours} hour{'s' if onsite_hours != 1 else ''}")
    lines.append(f"  - Total: {total_hours} hours")
    lines.append("")

    # Add standard deliverables
    lines.append("Deliverables:")
    lines.append("  - Network documentation and diagrams")
    lines.append("  - Device configuration backups")
    lines.append("  - Credentials documentation (secure)")

    return '\n'.join(lines)


def update_engineering_sow_tab(sow_file, summary_text):
    """Update the Engineering SOW tab with the summary"""
    if not openpyxl:
        print("Error: openpyxl not available")
        return False

    wb = openpyxl.load_workbook(sow_file)

    if 'Engineering SOW' not in wb.sheetnames:
        print("Error: Engineering SOW sheet not found")
        return False

    ws = wb['Engineering SOW']

    # B13 is the merged cell for Scope of Work content (B13:F22)
    ws['B13'] = summary_text

    # Save
    wb.save(sow_file)
    print(f"Updated {sow_file}")
    return True


def create_ecc_project_tasks(customer_name, equipment, sp_client=None):
    """Create project and tasks in Engineering Command Center via SharePoint Lists"""

    # This would interact with the SharePoint Lists that back the ECC
    # For now, generate the task list that should be created

    tasks = []

    # Calculate hours for tasks
    num_sites = len([loc for loc, items in equipment.get('locations', {}).items() if items])
    num_firewalls = sum(
        len([i for i in items if 'firewall' in i['description'].lower() or 'sonicwall' in i['manufacturer'].lower()])
        for items in equipment.get('locations', {}).values()
    )
    has_other_devices = any(
        any(i for i in items if 'firewall' not in i['description'].lower() and 'sonicwall' not in i['manufacturer'].lower())
        for items in equipment.get('locations', {}).values()
    )

    # Phase 1: Pre-Installation (Planning tasks - no billable hours)
    tasks.append({
        'title': 'Review network scan report and checklist',
        'phase': 'Planning',
        'status': 'To Do',
        'priority': 'High',
        'estimated_hours': 0
    })
    tasks.append({
        'title': 'Order equipment from vendors',
        'phase': 'Planning',
        'status': 'To Do',
        'priority': 'High',
        'estimated_hours': 0
    })
    tasks.append({
        'title': 'Schedule installation date with customer',
        'phase': 'Planning',
        'status': 'To Do',
        'priority': 'High',
        'estimated_hours': 0
    })

    # Network Discovery task - 1 hour per site
    tasks.append({
        'title': f'Network Discovery ({num_sites} site{"s" if num_sites != 1 else ""})',
        'phase': 'Discovery',
        'status': 'To Do',
        'priority': 'High',
        'estimated_hours': num_sites
    })

    # Phase 2: Equipment per location
    for location, items in equipment.get('locations', {}).items():
        firewalls = [i for i in items if 'firewall' in i['description'].lower() or 'sonicwall' in i['manufacturer'].lower()]
        switches = [i for i in items if 'switch' in i['description'].lower()]
        aps = [i for i in items if 'ap' in i['description'].lower() or 'u7' in i['model'].lower()]

        loc_prefix = location.replace(' ', '_')[:20]

        # Firewall tasks - 1 hour per firewall
        for fw in firewalls:
            tasks.append({
                'title': f'Configure firewall - {location}',
                'phase': 'Execution',
                'status': 'To Do',
                'priority': 'High',
                'estimated_hours': 1
            })

        # Other device config - combined into one task per location
        if switches or aps:
            other_items = []
            if switches:
                other_items.append(f"{len(switches)} switch(es)")
            if aps:
                other_items.append(f"{sum(i.get('quantity', 1) for i in aps)} AP(s)")
            tasks.append({
                'title': f'Configure {", ".join(other_items)} - {location}',
                'phase': 'Execution',
                'status': 'To Do',
                'priority': 'Medium',
                'estimated_hours': 1 if (switches or aps) else 0
            })

    # Onsite Install/Verification - 1 hour per site
    tasks.append({
        'title': f'Onsite Install and Verification ({num_sites} site{"s" if num_sites != 1 else ""})',
        'phase': 'Testing',
        'status': 'To Do',
        'priority': 'High',
        'estimated_hours': num_sites
    })

    # Phase 4: Documentation (no billable hours - included in project)
    tasks.append({
        'title': 'Create network diagrams',
        'phase': 'Documentation',
        'status': 'To Do',
        'priority': 'Medium',
        'estimated_hours': 0
    })
    tasks.append({
        'title': 'Document device configurations',
        'phase': 'Documentation',
        'status': 'To Do',
        'priority': 'Medium',
        'estimated_hours': 0
    })
    tasks.append({
        'title': 'Update credentials in Keeper',
        'phase': 'Documentation',
        'status': 'To Do',
        'priority': 'High',
        'estimated_hours': 0
    })

    # Phase 5: Handoff
    tasks.append({
        'title': 'Get customer sign-off',
        'phase': 'Handoff',
        'status': 'To Do',
        'priority': 'High',
        'estimated_hours': 0
    })

    return tasks


def print_project_summary(customer_name, equipment, tasks, sow_summary):
    """Print summary of what will be created"""
    print("\n" + "=" * 60)
    print(f"NEW SITE INSTALL: {customer_name}")
    print("=" * 60)

    print("\n--- EQUIPMENT BY LOCATION ---")
    for location, items in equipment.get('locations', {}).items():
        print(f"\n{location}:")
        for item in items:
            qty = item['quantity']
            desc = item['description']
            mfr = item['manufacturer']
            model = item['model']
            print(f"  - {qty}x {mfr} {model} ({desc})" if mfr else f"  - {qty}x {desc}")

    print("\n--- ENGINEERING SOW SUMMARY ---")
    print(sow_summary[:500] + "..." if len(sow_summary) > 500 else sow_summary)

    total_hours = sum(t.get('estimated_hours', 0) for t in tasks)
    print(f"\n--- ECC TASKS ({len(tasks)} tasks, {total_hours} billable hours) ---")
    for i, task in enumerate(tasks[:10], 1):
        hours = task.get('estimated_hours', 0)
        hours_str = f" ({hours}h)" if hours > 0 else ""
        print(f"  {i}. [{task['phase']}] {task['title']}{hours_str}")
    if len(tasks) > 10:
        print(f"  ... and {len(tasks) - 10} more tasks")


def main():
    parser = argparse.ArgumentParser(description='New Site Install Automation')
    parser.add_argument('--customer', '-c', required=True, help='Customer name or alias (e.g., SFWB)')
    parser.add_argument('--generate-sow', action='store_true', help='Generate and update Engineering SOW')
    parser.add_argument('--create-project', action='store_true', help='Create ECC project and tasks')
    parser.add_argument('--upload-scans', action='store_true', help='Upload scan files to SharePoint Technical Docs/Scans')
    parser.add_argument('--generate-checklist', action='store_true', help='Generate network checklist with auto-filled data from scans/SOW')
    parser.add_argument('--allocate-subnet', action='store_true', help='Allocate a new subnet block from the Obera subnet pool')
    parser.add_argument('--all', action='store_true', help='Run all automation steps')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be done without making changes')
    args = parser.parse_args()

    customer_name = resolve_customer_name(args.customer)
    print(f"Customer: {customer_name}")

    # Find customer folder
    customer_folder = find_customer_folder(args.customer)
    if not customer_folder:
        print(f"Error: Customer folder not found for '{args.customer}'")
        print(f"Looking in: {OBERACONNECT_TECH}")
        return

    print(f"Found folder: {customer_folder}")

    # Find SOW file
    sow_file = find_sow_file(customer_folder)
    if not sow_file:
        print(f"Warning: SOW file not found in {customer_folder}")
    else:
        print(f"Found SOW: {sow_file.name}")

    # Extract equipment from Pricing SOW
    equipment = {'locations': {}}
    sow_summary = ""

    if sow_file and openpyxl:
        wb = openpyxl.load_workbook(sow_file)
        equipment = extract_equipment_from_pricing(wb)
        sow_summary = generate_engineering_sow_summary(equipment, customer_name)

    # Generate tasks
    tasks = create_ecc_project_tasks(customer_name, equipment)

    # Print summary
    print_project_summary(customer_name, equipment, tasks, sow_summary)

    if args.dry_run:
        print("\n[DRY RUN] No changes made.")
        return

    # Execute requested actions
    if args.generate_sow or args.all:
        if sow_file and sow_summary:
            print("\n--- Updating Engineering SOW tab ---")
            if update_engineering_sow_tab(sow_file, sow_summary):
                print("Engineering SOW updated successfully")

                # Upload to SharePoint
                sp = SharePointClient()
                if sp.authenticate() and sp.get_site_and_drive():
                    folder_path = f"{customer_name}/Statements of Work (SOWs)"
                    with open(sow_file, 'rb') as f:
                        content = f.read()
                    success, result = sp.upload_file(folder_path, sow_file.name, content)
                    if success:
                        print(f"Uploaded to SharePoint: {folder_path}/{sow_file.name}")
                    else:
                        print("Warning: Failed to upload to SharePoint")

    if args.create_project or args.all:
        print("\n--- Creating ECC Project ---")
        print("Project creation would add to SharePoint Lists:")
        print(f"  Project: {customer_name} - Network Installation")
        print(f"  Tasks: {len(tasks)} tasks across {len(set(t['phase'] for t in tasks))} phases")
        print("\n[Note: Full ECC integration requires SharePoint List API calls]")

        # Save tasks to JSON for manual import or API integration
        tasks_file = OUTPUT_DIR / f"ecc_tasks_{customer_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.json"
        with open(tasks_file, 'w') as f:
            json.dump({
                'customer': customer_name,
                'project': f"{customer_name} - Network Installation",
                'created': datetime.now().isoformat(),
                'tasks': tasks
            }, f, indent=2)
        print(f"Tasks saved to: {tasks_file}")

    if args.upload_scans or args.all:
        print("\n--- Uploading Scan Files ---")
        scan_files = find_scan_files(args.customer)

        if not scan_files:
            print(f"No scan files found for {customer_name}")
        else:
            print(f"Found {len(scan_files)} scan file(s):")
            for f in scan_files:
                print(f"  - {f.name}")

            sp = SharePointClient()
            if sp.authenticate() and sp.get_site_and_drive():
                uploaded, failed = upload_scans_to_sharepoint(
                    customer_name, scan_files, sp, dry_run=args.dry_run
                )
                print(f"\nUploaded: {len(uploaded)}, Failed: {len(failed)}")
            else:
                print("Error: Could not authenticate to SharePoint")

    if args.allocate_subnet or args.all:
        print("\n--- Subnet Allocation ---")

        # Check if already allocated
        existing = get_customer_subnet_allocation(customer_name)
        if existing:
            print(f"Customer already has subnet allocated:")
            print(f"  Subnet: {existing['assigned_subnet']}")
            print(f"  Gateway: {existing['assigned_gateway']}")
            print(f"  Block: {existing['allocated_block']}")
        else:
            if SUBNET_ALLOCATOR_AVAILABLE:
                next_subnet = get_next_available_subnet()
                print(f"Next available subnet: {next_subnet['subnet'] if next_subnet else 'N/A'}")

                if args.dry_run:
                    print("[DRY RUN] Would allocate subnet block")
                else:
                    allocation = allocate_new_subnet(customer_name)
                    if allocation:
                        print(f"Allocated new subnet block:")
                        print(f"  Customer ID: {allocation['customer_id']}")
                        print(f"  Block: {allocation['allocated_block']}")
                        print(f"  Subnet: {allocation['assigned_subnet']}")
                        print(f"  Gateway: {allocation['assigned_gateway']}")
                        print(f"  DHCP Range: {allocation['dhcp_range']}")
                        print(f"  Registry: {SUBNET_REGISTRY_PATH}")
                    else:
                        print("Failed to allocate subnet")
            else:
                print("Error: Subnet allocator not available")
                print(f"  Expected at: {SUBNET_ALLOCATOR_PATH}")

    if args.generate_checklist or args.all:
        print("\n--- Generating Network Checklist ---")

        # Gather all available data from scans and SOW
        checklist_data = gather_checklist_data(customer_name)

        # Generate HTML
        html_content = generate_checklist_html(checklist_data)

        # Save locally
        safe_name = customer_name.replace(' ', '_')
        filename = f"Network_Checklist_{safe_name}_{checklist_data['date']}.html"
        local_path = OUTPUT_DIR / filename
        local_path.write_text(html_content)
        print(f"\nSaved locally: {local_path}")

        # Upload to SharePoint
        sp = SharePointClient()
        if sp.authenticate() and sp.get_site_and_drive():
            folder_path = f"{customer_name}/Technical Docs"
            success, result = sp.upload_file(folder_path, filename, html_content.encode('utf-8'))
            if success:
                print(f"Uploaded to SharePoint: {folder_path}/{filename}")
            else:
                print("Warning: Failed to upload to SharePoint")

        # Summary of what was auto-filled vs manual
        auto_count = len(checklist_data.get('auto_filled_fields', []))
        total_fields = 25  # Approximate number of info fields
        print(f"\n  Auto-filled: {auto_count}/{total_fields} fields ({int(auto_count/total_fields*100)}%)")
        print(f"  Manual entry still required: {total_fields - auto_count} fields")

    print("\n" + "=" * 60)
    print("AUTOMATION COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
